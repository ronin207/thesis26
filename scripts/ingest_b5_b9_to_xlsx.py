#!/usr/bin/env python3
"""Ingest B5 (Griffin breakdown) + B9 (PP3 policy) JSONL into bench_results.xlsx.

Adds/refreshes three sheets, preserving every existing sheet:
  * "B5 Griffin Breakdown (M5 Pro)"  — one row per phase from breakdown_phase samples.
  * "Griffin Coarse (M5 Pro)"        — REBUILT from the fresh B5 JSONL summaries.
  * "PP3 Policy (M5 Pro)"            — 3-row D2-style table from B9 pp3_policy summaries.

Source JSONLs (from scripts/run_benchmarks.sh --snark --only 'B[59]'):
  results/bench_run_20260416_112256/bench_runner_B5_griffin.jsonl
  results/bench_run_20260416_112256/bench_runner_B9_pp3_policy.jsonl

Re-runnable: the three target sheets are deleted+recreated; everything else is
untouched. Output is written back to results/bench_results.xlsx in place.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
RUN_DIR = ROOT / "results" / "bench_run_20260416_112256"
B5_PATH = RUN_DIR / "bench_runner_B5_griffin.jsonl"
B9_PATH = RUN_DIR / "bench_runner_B9_pp3_policy.jsonl"
XLSX = ROOT / "results" / "bench_results.xlsx"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFEEEEEE")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, italic=True, color="FF444444")


def load_jsonl(path: Path) -> list[dict]:
    records = []
    with path.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            records.append(json.loads(line))
    return records


def drop_sheet(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def autosize(ws, min_w: int = 10, max_w: int = 46) -> None:
    for col in ws.columns:
        length = min_w
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = max(length, min(max_w, len(str(v)) + 2))
        ws.column_dimensions[letter].width = length


def write_header_row(ws, row_idx: int, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left")


# ──────────────────────────────────────────────────────────────────────────────
# B5: Griffin component breakdown sheet (NEW)
# ──────────────────────────────────────────────────────────────────────────────

def build_b5_breakdown_sheet(wb, b5: list[dict]) -> None:
    name = "B5 Griffin Breakdown (M5 Pro)"
    drop_sheet(wb, name)
    ws = wb.create_sheet(name)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(
        row=1, column=1,
        value=(
            f"Generated: {ts} — M5 Pro  (B5 breakdown_phase: Griffin cost per phase "
            f"inside build_loquat_r1cs, k=2, m=64, rev_depth=20)"
        ),
    ).font = TITLE_FONT

    headers = [
        "phase",
        "delta_constraints",
        "delta_variables",
        "fraction_of_loquat_verify",
        "fraction_of_full_credential",
        "notes",
    ]
    write_header_row(ws, 2, headers)

    # Source phase ordering from JSONL (stable across the single run).
    phases = [r for r in b5 if r.get("type") == "sample" and r.get("variant") == "breakdown_phase"]
    if not phases:
        raise RuntimeError("no breakdown_phase samples found in B5 JSONL")

    phase_notes = {
        "message commitment":   "Griffin hash of signed message → 32-byte digest",
        "t-values allocated":   "witness allocation only, no constraints",
        "quadratic residuosity": "Legendre-symbol style QR check per Loquat §4.3",
        "transcript binding":   "Fiat-Shamir: Griffin re-absorbs every prover message",
        "sumcheck":             "per-round arithmetic (no hashing)",
        "LDT queries":          "Griffin Merkle auth paths for every FRI query",
    }

    row = 3
    total_loquat = 0
    total_full_ref = 0
    for rec in phases:
        m = rec["metrics"]
        dc, dv = int(m["delta_constraints"]), int(m["delta_variables"])
        total_loquat += dc
        total_full_ref = int(m["full_constraint_ref"])
        ws.cell(row=row, column=1, value=rec["config_key"])
        ws.cell(row=row, column=2, value=dc)
        ws.cell(row=row, column=3, value=dv)
        ws.cell(row=row, column=4, value=float(m["fraction_of_loquat"]))
        ws.cell(row=row, column=5, value=float(m["fraction_of_full"]))
        ws.cell(row=row, column=6, value=phase_notes.get(rec["config_key"], ""))
        row += 1

    # Totals row.
    ws.cell(row=row, column=1, value="TOTAL (Loquat sig-verify)").font = HEADER_FONT
    ws.cell(row=row, column=2, value=total_loquat).font = HEADER_FONT
    ws.cell(row=row, column=4, value=1.0).font = HEADER_FONT
    ws.cell(row=row, column=5, value=total_loquat / total_full_ref if total_full_ref else None)
    ws.cell(row=row, column=6, value="sum of phase Δc above")
    row += 1
    ws.cell(row=row, column=1, value="TOTAL (Full credential, k=2)").font = HEADER_FONT
    ws.cell(row=row, column=2, value=total_full_ref).font = HEADER_FONT
    ws.cell(row=row, column=5, value=1.0).font = HEADER_FONT
    ws.cell(row=row, column=6, value="from full_credential sample (B5)")
    row += 2

    # Griffin vs non-Griffin aggregate.
    griffin_phases = {"message commitment", "transcript binding", "LDT queries"}
    griffin_delta = sum(
        int(r["metrics"]["delta_constraints"])
        for r in phases if r["config_key"] in griffin_phases
    )
    other_delta = total_loquat - griffin_delta

    ws.cell(row=row, column=1, value="Derived: Griffin vs non-Griffin share").font = TITLE_FONT
    row += 1
    for label, val in [
        ("Griffin-heavy phases (msg+transcript+LDT) Δc", griffin_delta),
        ("Non-Griffin phases Δc", other_delta),
        ("Griffin share of Loquat sig-verify", f"{griffin_delta/total_loquat*100:.2f}%"),
        ("Griffin share of full credential", f"{griffin_delta/total_full_ref*100:.2f}%"),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1

    # Format fraction columns as percentages for readability.
    for r in range(3, 3 + len(phases) + 2):
        for col in (4, 5):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
    autosize(ws)


# ──────────────────────────────────────────────────────────────────────────────
# Griffin Coarse sheet (REBUILT from the new B5 JSONL)
# ──────────────────────────────────────────────────────────────────────────────

def build_griffin_coarse_sheet(wb, b5: list[dict]) -> None:
    name = "Griffin Coarse (M5 Pro)"
    drop_sheet(wb, name)
    ws = wb.create_sheet(name)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(
        row=1, column=1,
        value=f"Generated: {ts} — M5 Pro  (B5 griffin: full_credential vs loquat_only_aurora)",
    ).font = TITLE_FONT

    headers = ["variant", "config_key", "metric", "mean", "std_dev", "median", "n_runs"]
    write_header_row(ws, 2, headers)

    order = [
        ("full_credential", "full_k=2_m=64_rev20",
         ["constraint_count", "prove_ms", "verify_ms", "proof_bytes"]),
        ("loquat_only_aurora", "loquat_sig_verify",
         ["prove_ms", "verify_ms", "proof_bytes"]),
    ]

    def find(variant: str, metric: str) -> dict | None:
        for r in b5:
            if (
                r.get("type") == "summary"
                and r.get("variant") == variant
                and r.get("metric") == metric
            ):
                return r
        return None

    row = 3
    full_prove_mean = None
    loquat_prove_mean = None
    for variant, cfg_key, metrics in order:
        for metric in metrics:
            s = find(variant, metric)
            ws.cell(row=row, column=1, value=variant)
            ws.cell(row=row, column=2, value=cfg_key)
            ws.cell(row=row, column=3, value=metric)
            if s is not None:
                ws.cell(row=row, column=4, value=round(float(s["mean"]), 2))
                ws.cell(row=row, column=5, value=round(float(s["std_dev"]), 2))
                ws.cell(row=row, column=6, value=round(float(s["median"]), 2))
                ws.cell(row=row, column=7, value=int(s["n_after_filter"]))
                if variant == "full_credential" and metric == "prove_ms":
                    full_prove_mean = float(s["mean"])
                if variant == "loquat_only_aurora" and metric == "prove_ms":
                    loquat_prove_mean = float(s["mean"])
            row += 1

    # Derived block.
    row += 1
    ws.cell(row=row, column=1, value="Derived: coarse infrastructure attribution").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="full_credential prove_ms")
    ws.cell(row=row, column=2, value=round(full_prove_mean, 1) if full_prove_mean else None)
    row += 1
    ws.cell(row=row, column=1, value="loquat_only prove_ms")
    ws.cell(row=row, column=2, value=round(loquat_prove_mean, 1) if loquat_prove_mean else None)
    row += 1
    if full_prove_mean and loquat_prove_mean:
        overhead_x = full_prove_mean / loquat_prove_mean
        overhead_pct = (1.0 - loquat_prove_mean / full_prove_mean) * 100
        ws.cell(row=row, column=1, value="infrastructure overhead (×)")
        ws.cell(row=row, column=2, value=round(overhead_x, 2))
        row += 1
        ws.cell(row=row, column=1, value="infrastructure share (%)")
        ws.cell(row=row, column=2, value=round(overhead_pct, 1))
    autosize(ws)


# ──────────────────────────────────────────────────────────────────────────────
# PP3 Policy sheet (NEW from B9)
# ──────────────────────────────────────────────────────────────────────────────

def build_pp3_policy_sheet(wb, b9: list[dict]) -> None:
    name = "PP3 Policy (M5 Pro)"
    drop_sheet(wb, name)
    ws = wb.create_sheet(name)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(
        row=1, column=1,
        value=(
            f"Generated: {ts} — M5 Pro  (B9 pp3_policy: prove/verify cost vs policy "
            f"predicate set; k=2, lr=0, rev_depth=20, 10 measured runs)"
        ),
    ).font = TITLE_FONT

    headers = [
        "policy",
        "constraint_count",
        "indexer_ms",
        "prove_ms",
        "verify_ms",
        "proof_bytes",
        "proof_KiB",
        "n_runs",
        "prove_std_dev",
        "verify_std_dev",
    ]
    write_header_row(ws, 2, headers)

    def find_summary(cfg_key: str, metric: str) -> dict | None:
        for r in b9:
            if (
                r.get("type") == "summary"
                and r.get("config_key") == cfg_key
                and r.get("metric") == metric
            ):
                return r
        return None

    def mean(cfg_key: str, metric: str) -> float | None:
        s = find_summary(cfg_key, metric)
        return float(s["mean"]) if s else None

    def stddev(cfg_key: str, metric: str) -> float | None:
        s = find_summary(cfg_key, metric)
        return float(s["std_dev"]) if s and s.get("std_dev") is not None else None

    def n_runs(cfg_key: str, metric: str) -> int | None:
        s = find_summary(cfg_key, metric)
        return int(s["n_after_filter"]) if s else None

    policies = [
        ("none",       "k=2_lr=0_rev20_pol=none"),
        ("gpa",        "k=2_lr=0_rev20_pol=gpa"),
        ("gpa_degree", "k=2_lr=0_rev20_pol=gpa_degree"),
    ]

    row = 3
    none_constraints = mean(policies[0][1], "constraint_count")
    none_prove = mean(policies[0][1], "prove_ms")
    none_verify = mean(policies[0][1], "verify_ms")
    none_bytes = mean(policies[0][1], "proof_bytes")
    for label, cfg_key in policies:
        cc = mean(cfg_key, "constraint_count")
        pr = mean(cfg_key, "prove_ms")
        vr = mean(cfg_key, "verify_ms")
        pb = mean(cfg_key, "proof_bytes")
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=int(cc) if cc is not None else None)
        ws.cell(row=row, column=3, value=round(mean(cfg_key, "indexer_ms"), 2))
        ws.cell(row=row, column=4, value=round(pr, 2) if pr is not None else None)
        ws.cell(row=row, column=5, value=round(vr, 2) if vr is not None else None)
        ws.cell(row=row, column=6, value=int(pb) if pb is not None else None)
        ws.cell(row=row, column=7, value=round(pb / 1024.0, 1) if pb is not None else None)
        ws.cell(row=row, column=8, value=n_runs(cfg_key, "prove_ms"))
        pr_sd = stddev(cfg_key, "prove_ms")
        vr_sd = stddev(cfg_key, "verify_ms")
        ws.cell(row=row, column=9, value=round(pr_sd, 2) if pr_sd is not None else None)
        ws.cell(row=row, column=10, value=round(vr_sd, 2) if vr_sd is not None else None)
        row += 1

    # Derived block: Δ vs pol=none.
    row += 1
    ws.cell(row=row, column=1, value="Derived: overhead vs pol=none").font = TITLE_FONT
    row += 1
    write_header_row(ws, row, [
        "policy", "Δ constraints", "Δ prove_ms", "Δ prove (%)",
        "Δ verify_ms", "Δ verify (%)", "Δ proof_bytes",
    ])
    row += 1
    for label, cfg_key in policies[1:]:
        cc = mean(cfg_key, "constraint_count")
        pr = mean(cfg_key, "prove_ms")
        vr = mean(cfg_key, "verify_ms")
        pb = mean(cfg_key, "proof_bytes")
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=int(cc - none_constraints))
        ws.cell(row=row, column=3, value=round(pr - none_prove, 2))
        ws.cell(row=row, column=4, value=round((pr - none_prove) / none_prove * 100, 3))
        ws.cell(row=row, column=5, value=round(vr - none_verify, 2))
        ws.cell(row=row, column=6, value=round((vr - none_verify) / none_verify * 100, 3))
        ws.cell(row=row, column=7, value=int(pb - none_bytes))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=(
        "Note: Δ proof_bytes is near zero because Aurora proof length is pinned to "
        "the next power of two (2^20 for 10^6 constraints); policy deltas of <300 "
        "constraints stay within the same power-of-two bucket."
    )).font = Font(italic=True, color="FF666666")
    autosize(ws)


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main() -> int:
    if not B5_PATH.exists():
        print(f"error: {B5_PATH} not found", file=sys.stderr)
        return 2
    if not B9_PATH.exists():
        print(f"error: {B9_PATH} not found", file=sys.stderr)
        return 2
    if not XLSX.exists():
        print(f"error: {XLSX} not found", file=sys.stderr)
        return 2

    b5 = load_jsonl(B5_PATH)
    b9 = load_jsonl(B9_PATH)
    print(f"[ingest] loaded {len(b5)} B5 records, {len(b9)} B9 records")

    wb = load_workbook(XLSX)
    before = list(wb.sheetnames)

    build_b5_breakdown_sheet(wb, b5)
    build_griffin_coarse_sheet(wb, b5)
    build_pp3_policy_sheet(wb, b9)

    wb.save(XLSX)

    after = list(wb.sheetnames)
    added = [s for s in after if s not in before]
    rebuilt = [s for s in after if s in before and s in (
        "Griffin Coarse (M5 Pro)", "B5 Griffin Breakdown (M5 Pro)", "PP3 Policy (M5 Pro)"
    )]
    print(f"[ingest] wrote {XLSX.relative_to(ROOT)}")
    print(f"  + added:    {added}")
    print(f"  + rebuilt:  {rebuilt}")
    print(f"  + total sheets: {len(after)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
