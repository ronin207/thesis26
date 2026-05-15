#!/usr/bin/env python3
"""
JSONL → Excel pipeline for vc-pqc thesis benchmarks.

Reads all bench_*.jsonl files from the results directory, picks the best
(most recent, non-anomalous) measurements for each suite, and writes a
fully-populated Excel workbook with charts.

Usage:
    python3 create_analysis.py [--output PATH] [--file FILE1 FILE2 ...]

If --file is omitted, all results/bench_*.jsonl files are used.
The M5 Pro run (bench_1776145971) takes precedence when available.
"""

import argparse
import json
import math
import os
import re
import statistics
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Optional openpyxl chart imports
# ---------------------------------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference, Series
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not found. Install with: pip install openpyxl", file=sys.stderr)

RESULTS_DIR = Path(__file__).parent

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
ALT_FILL     = PatternFill('solid', fgColor='D6E4F0')
WHITE_FILL   = PatternFill('solid', fgColor='FFFFFF')
SUMMARY_FILL = PatternFill('solid', fgColor='E2EFDA')
TS_FONT      = Font(name='Arial', size=8, color='999999', italic=True)
DATA_FONT    = Font(name='Arial', size=10)
BOLD_FONT    = Font(name='Arial', size=10, bold=True)
THIN_BORDER  = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)

# ---------------------------------------------------------------------------
# JSONL loading
# ---------------------------------------------------------------------------

def load_jsonl(path):
    """Return list of parsed JSON objects from a JSONL file."""
    records = []
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return records


def collect_records(jsonl_files):
    """
    Load all JSONL files and return (samples, summaries, headers).
    samples   : list of {type:'sample', ...} records
    summaries : list of {type:'summary', ...} records
    headers   : list of {type:'header', ...} records (one per file)
    """
    samples, summaries, headers = [], [], []
    for p in jsonl_files:
        for rec in load_jsonl(p):
            t = rec.get('type', '')
            if t == 'sample':
                rec['_source_file'] = str(p)
                samples.append(rec)
            elif t == 'summary':
                rec['_source_file'] = str(p)
                summaries.append(rec)
            elif t == 'header':
                rec['_source_file'] = str(p)
                headers.append(rec)
    return samples, summaries, headers


def measured(samples):
    """Filter to non-warmup samples only."""
    return [s for s in samples if not s.get('is_warmup', False)]


# Files known to contain toy-aurora data (proof_bytes < 1 MB for B4/B5/B7).
# These are from the feature/succinct-proof-measurement branch POC.
_TOY_AURORA_FILES = {
    'bench_post_aurora_fix.jsonl',
    'bench_b1_b2_rerun.jsonl',
    'bench_b2_b5_rerun.jsonl',
    'bench_b2_rerun2.jsonl',
}

# For suites that use libiop Aurora, reject records where proof_bytes is
# suspiciously small (< 1 MB means toy POC, not real libiop 33 MB output).
_LIBIOP_SUITES = {'B4', 'B5', 'B7'}


def is_libiop_record(sample):
    """Return True if a sample appears to be from a real libiop Aurora run."""
    fname = Path(sample.get('_source_file', '')).name
    if fname in _TOY_AURORA_FILES:
        return False
    suite = sample.get('suite', '')
    if suite in _LIBIOP_SUITES:
        pb = (sample.get('metrics') or {}).get('proof_bytes', 0)
        # libiop Aurora proofs are ≥ 8 MB; toy POC ≤ 50 KB
        if pb and pb < 1_000_000:
            return False
    return True


def canonical(samples):
    """
    Return samples filtered to canonical (real-libiop) runs only.
    For each (suite, variant, config_key), prefer the M5 Pro run if present;
    otherwise keep all canonical M2 records.
    """
    real = [s for s in samples if is_libiop_record(s)]

    # Determine if M5 Pro data exists (hostname contains M5 or file is bench_1776*)
    m5_files = {Path(s['_source_file']).name for s in real
                if '1776' in s.get('_source_file', '')}
    if not m5_files:
        return real  # no M5 data yet; return all canonical records

    # Prefer M5 Pro file for suites it covers; keep M2 for suites it doesn't
    m5_suites = {s.get('suite') for s in real
                 if Path(s['_source_file']).name in m5_files}

    out = []
    for s in real:
        fname = Path(s['_source_file']).name
        suite = s.get('suite')
        if suite in m5_suites:
            if fname in m5_files:
                out.append(s)
        else:
            out.append(s)
    return out


def suite_samples(samples, suite, variant=None):
    """Filter samples by suite and optionally variant."""
    out = [s for s in samples if s.get('suite') == suite]
    if variant:
        out = [s for s in out if s.get('variant') == variant]
    return measured(out)


def suite_summaries(summaries, suite, variant=None, metric=None):
    """Filter summaries by suite, variant, metric."""
    out = [s for s in summaries if s.get('suite') == suite]
    if variant:
        out = [s for s in out if s.get('variant') == variant]
    if metric:
        out = [s for s in out if s.get('metric') == metric]
    return out


def get_phases(sample, key):
    return sample.get('phases', {}).get(key)


def get_metrics(sample, key):
    return sample.get('metrics', {}).get(key)


def iqr_filter(values):
    """Remove outliers beyond 1.5×IQR from Q1/Q3."""
    if len(values) < 4:
        return values
    sv = sorted(values)
    n = len(sv)
    q1 = sv[n // 4]
    q3 = sv[(3 * n) // 4]
    iqr = q3 - q1
    lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    return [v for v in values if lo <= v <= hi]


def mean_std(values):
    if not values:
        return None, None
    v = iqr_filter(values)
    if not v:
        v = values
    m = statistics.mean(v)
    s = statistics.stdev(v) if len(v) > 1 else 0.0
    return m, s

# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------

def setup_sheet(ws, title, headers, col_widths=None):
    ws.title = title
    ws['A1'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A1'].font = TS_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = THIN_BORDER
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'


def style_rows(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if (r - start_row) % 2 == 0 else WHITE_FILL
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER


def style_summary(ws, row, num_cols, label=None):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_FONT
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER
    if label:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT


def write_val(ws, row, col, val, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    if fmt:
        c.number_format = fmt
    return c

# ---------------------------------------------------------------------------
# Sheet 1: B7 PP2 Aurora (D2) — main performance table
# ---------------------------------------------------------------------------

def build_sheet_pp2(wb, samples, headers_info):
    """PP2 Aurora: k=2, lr=0, rev_depth=20."""
    ws = wb.active
    cols = ['Run#', 'prove_s', 'verify_s', 'indexer_s', 'proof_verify_s',
            'constraint_count', 'proof_bytes', 'proof_MiB', 'signature_bytes', 'source_file']
    widths = [6, 12, 12, 12, 14, 18, 14, 12, 16, 30]
    setup_sheet(ws, 'PP2 Aurora (D2)', cols, widths)

    recs = suite_samples(samples, 'B7', 'pp2_aurora')
    if not recs:
        ws.cell(row=3, column=1, value='NO DATA').font = BOLD_FONT
        return ws

    data_start = 3
    for i, s in enumerate(recs):
        r = data_start + i
        prove   = (get_phases(s, 'prove_ms') or s.get('wall_ms', 0)) / 1000
        verify  = (get_phases(s, 'verify_ms') or 0) / 1000
        indexer = (get_phases(s, 'indexer_ms') or 0) / 1000
        pv      = (get_phases(s, 'proof_verify_ms') or 0) / 1000
        cc      = get_metrics(s, 'constraint_count') or 0
        pb      = get_metrics(s, 'proof_bytes') or 0
        sb      = get_metrics(s, 'signature_bytes') or 0
        src     = Path(s.get('_source_file', '')).name

        write_val(ws, r, 1, s.get('run', i))
        write_val(ws, r, 2, prove,   '0.000')
        write_val(ws, r, 3, verify,  '0.000')
        write_val(ws, r, 4, indexer, '0.000')
        write_val(ws, r, 5, pv,      '0.000')
        write_val(ws, r, 6, cc,      '#,##0')
        write_val(ws, r, 7, pb,      '#,##0')
        write_val(ws, r, 8, pb / (1024**2) if pb else None, '0.00')
        write_val(ws, r, 9, sb,      '#,##0')
        write_val(ws, r, 10, src)

    data_end = data_start + len(recs) - 1
    style_rows(ws, data_start, data_end, len(cols))

    # Summary rows
    stat_start = data_end + 2
    prove_vals  = [(get_phases(s, 'prove_ms') or s.get('wall_ms', 0)) / 1000 for s in recs]
    verify_vals = [(get_phases(s, 'verify_ms') or 0) / 1000 for s in recs]
    pb_vals     = [get_metrics(s, 'proof_bytes') or 0 for s in recs]
    cc_vals     = [get_metrics(s, 'constraint_count') or 0 for s in recs]

    for label, fn in [('MEAN', statistics.mean), ('STDEV', lambda v: statistics.stdev(v) if len(v)>1 else 0)]:
        r = stat_start
        stat_start += 1
        style_summary(ws, r, len(cols), label)
        write_val(ws, r, 2, fn(prove_vals),  '0.000')
        write_val(ws, r, 3, fn(verify_vals), '0.000')
        write_val(ws, r, 6, fn(cc_vals),     '#,##0')
        write_val(ws, r, 7, fn(pb_vals),     '#,##0')
        write_val(ws, r, 8, fn(pb_vals) / (1024**2), '0.00')

    # Annotate header info
    if headers_info:
        h = headers_info[0]
        ws['A1'] = (
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
            f"CPU: {h.get('cpu','')} | OS: {h.get('os','')} | n={len(recs)} non-warmup runs"
        )
        ws['A1'].font = TS_FONT
    return ws

# ---------------------------------------------------------------------------
# Sheet 2: B7 PP3 Policy (D2)
# ---------------------------------------------------------------------------

def build_sheet_pp3(wb, samples):
    ws = wb.create_sheet()
    cols = ['Run#', 'variant', 'prove_s', 'verify_s', 'constraint_count',
            'proof_bytes', 'proof_MiB', 'source_file']
    widths = [6, 20, 12, 12, 18, 14, 12, 30]
    setup_sheet(ws, 'PP3 Policy (D2)', cols, widths)

    variants = ['pp3_aurora', 'pp2_combined']
    all_recs = []
    for v in variants:
        for s in suite_samples(samples, 'B7', v):
            s['_variant_label'] = v
            all_recs.append(s)

    if not all_recs:
        ws.cell(row=3, column=1, value='NO DATA').font = BOLD_FONT
        return ws

    data_start = 3
    for i, s in enumerate(all_recs):
        r = data_start + i
        prove  = (get_phases(s, 'prove_ms') or s.get('wall_ms', 0)) / 1000
        verify = (get_phases(s, 'verify_ms') or 0) / 1000
        cc     = get_metrics(s, 'constraint_count') or 0
        pb     = get_metrics(s, 'proof_bytes') or 0

        write_val(ws, r, 1, s.get('run', i))
        write_val(ws, r, 2, s['_variant_label'])
        write_val(ws, r, 3, prove,  '0.000')
        write_val(ws, r, 4, verify, '0.000')
        write_val(ws, r, 5, cc,     '#,##0')
        write_val(ws, r, 6, pb,     '#,##0')
        write_val(ws, r, 7, pb / (1024**2) if pb else None, '0.00')
        write_val(ws, r, 8, Path(s.get('_source_file', '')).name)

    data_end = data_start + len(all_recs) - 1
    style_rows(ws, data_start, data_end, len(cols))

    # Per-variant summaries
    stat_row = data_end + 2
    for v in variants:
        vrecs = [s for s in all_recs if s['_variant_label'] == v]
        if not vrecs:
            continue
        prove_vals  = [(get_phases(s, 'prove_ms') or s.get('wall_ms', 0)) / 1000 for s in vrecs]
        verify_vals = [(get_phases(s, 'verify_ms') or 0) / 1000 for s in vrecs]
        cc_vals     = [get_metrics(s, 'constraint_count') or 0 for s in vrecs]
        pb_vals     = [get_metrics(s, 'proof_bytes') or 0 for s in vrecs]

        for label, fn in [('MEAN', statistics.mean), ('STDEV', lambda vals: statistics.stdev(vals) if len(vals)>1 else 0)]:
            style_summary(ws, stat_row, len(cols), f'{label} ({v})')
            write_val(ws, stat_row, 2, v)
            write_val(ws, stat_row, 3, fn(prove_vals),  '0.000')
            write_val(ws, stat_row, 4, fn(verify_vals), '0.000')
            write_val(ws, stat_row, 5, fn(cc_vals),     '#,##0')
            write_val(ws, stat_row, 6, fn(pb_vals),     '#,##0')
            write_val(ws, stat_row, 7, fn(pb_vals) / (1024**2), '0.00')
            stat_row += 1

    return ws

# ---------------------------------------------------------------------------
# Sheet 3: Constraint Scaling (B3)
# ---------------------------------------------------------------------------

def build_sheet_scaling(wb, samples, summaries):
    ws = wb.create_sheet()
    cols = ['k', 'rev_depth', 'policy', 'constraint_count', 'prove_s', 'verify_s',
            'proof_bytes', 'proof_MiB', 'config_key']
    widths = [6, 12, 16, 18, 12, 12, 14, 12, 30]
    setup_sheet(ws, 'Constraint Scaling (B3)', cols, widths)

    recs = suite_samples(samples, 'B3', 'constraints_only')
    if not recs:
        ws.cell(row=3, column=1, value='NO DATA').font = BOLD_FONT
        return ws

    # Sort by k, rev_depth, policy
    def sort_key(s):
        cfg = s.get('config', {})
        pol_order = {'none': 0, 'gpa': 1, 'gpa_degree': 2}
        return (cfg.get('k', 0), cfg.get('rev_depth', 0),
                pol_order.get(cfg.get('policy', 'none'), 9))

    recs.sort(key=sort_key)

    data_start = 3
    for i, s in enumerate(recs):
        r = data_start + i
        cfg = s.get('config', {})
        prove  = (get_phases(s, 'prove_ms') or 0) / 1000
        verify = (get_phases(s, 'verify_ms') or 0) / 1000
        cc     = get_metrics(s, 'constraint_count') or cfg.get('constraint_count', 0)
        pb     = get_metrics(s, 'proof_bytes') or 0

        write_val(ws, r, 1, cfg.get('k'))
        write_val(ws, r, 2, cfg.get('rev_depth'))
        write_val(ws, r, 3, cfg.get('policy', 'none'))
        write_val(ws, r, 4, cc,     '#,##0')
        write_val(ws, r, 5, prove,  '0.000')
        write_val(ws, r, 6, verify, '0.000')
        write_val(ws, r, 7, pb,     '#,##0')
        write_val(ws, r, 8, pb / (1024**2) if pb else None, '0.00')
        write_val(ws, r, 9, s.get('config_key', ''))

    data_end = data_start + len(recs) - 1
    style_rows(ws, data_start, data_end, len(cols))

    # Linear regression: N_C(k) = a + b*k (none policy, rev_depth=20)
    none20 = [(s.get('config', {}).get('k', 0),
               get_metrics(s, 'constraint_count') or s.get('config', {}).get('constraint_count', 0))
              for s in recs
              if s.get('config', {}).get('policy') == 'none'
              and s.get('config', {}).get('rev_depth') == 20]
    if len(none20) >= 2:
        xs = [x for x, _ in none20]
        ys = [y for _, y in none20]
        n  = len(xs)
        sx = sum(xs); sy = sum(ys)
        sxy = sum(x*y for x,y in zip(xs,ys))
        sxx = sum(x*x for x in xs)
        denom = n * sxx - sx**2
        b = (n * sxy - sx * sy) / denom if denom else 0
        a = (sy - b * sx) / n
        ss_res = sum((y - (a + b*x))**2 for x, y in zip(xs, ys))
        ss_tot = sum((y - sy/n)**2 for y in ys)
        r2 = 1 - ss_res/ss_tot if ss_tot else 1.0

        reg_row = data_end + 3
        ws.cell(row=reg_row, column=1,
                value='LINEAR REGRESSION (policy=none, rev_depth=20): N_C(k) = a + b·k'
                ).font = Font(name='Arial', bold=True, size=11, color='1F4E79')
        ws.merge_cells(start_row=reg_row, start_column=1, end_row=reg_row, end_column=5)

        for label, val, fmt in [
            ('Slope b (marginal cost per k)', b, '#,##0.0'),
            ('Intercept a (fixed cost)',       a, '#,##0.0'),
            ('R²',                             r2,'0.000000'),
        ]:
            reg_row += 1
            style_summary(ws, reg_row, 3, label)
            write_val(ws, reg_row, 2, val, fmt)

    return ws

# ---------------------------------------------------------------------------
# Sheet 4: Griffin Breakdown (B5)
# ---------------------------------------------------------------------------

def build_sheet_griffin(wb, samples):
    ws = wb.create_sheet()
    cols = ['Component / Config', 'constraint_count', '% of full ShowVer',
            'prove_s', 'verify_s', 'proof_bytes', 'Notes']
    widths = [30, 18, 18, 12, 12, 14, 36]
    setup_sheet(ws, 'Griffin Breakdown (B5)', cols, widths)

    def get_mean(variant):
        recs = suite_samples(samples, 'B5', variant)
        if not recs:
            return {}
        cc_v     = [get_metrics(s, 'constraint_count') or 0 for s in recs]
        prove_v  = [(get_phases(s, 'prove_ms') or 0) / 1000 for s in recs]
        verify_v = [(get_phases(s, 'verify_ms') or 0) / 1000 for s in recs]
        pb_v     = [get_metrics(s, 'proof_bytes') or 0 for s in recs]
        return {
            'cc': statistics.mean(cc_v) if cc_v else None,
            'prove': statistics.mean(prove_v) if prove_v else None,
            'verify': statistics.mean(verify_v) if verify_v else None,
            'pb': statistics.mean(pb_v) if pb_v else None,
            'n': len(recs),
        }

    full    = get_mean('full_credential')
    griffin = get_mean('griffin_only')
    merkle  = get_mean('merkle_only')

    full_cc = full.get('cc') or 1  # avoid division by zero

    rows_data = [
        ('Full ShowVer (k=2, m=64, rev=20)', full.get('cc'), 1.0,
         full.get('prove'), full.get('verify'), full.get('pb'),
         f'n={full.get("n",0)}; reference for % calculations'),
        ('Griffin hash only (isolated)',      griffin.get('cc'), (griffin.get('cc') or 0) / full_cc,
         griffin.get('prove'), griffin.get('verify'), griffin.get('pb'),
         f'n={griffin.get("n",0)}; B5 griffin_only variant'),
        ('Merkle path (revocation tree)',     merkle.get('cc'), (merkle.get('cc') or 0) / full_cc,
         merkle.get('prove'), merkle.get('verify'), merkle.get('pb'),
         f'n={merkle.get("n",0)}; depth=20'),
        ('Loquat+infrastructure (derived)',
         (full.get('cc') or 0) - (griffin.get('cc') or 0) - (merkle.get('cc') or 0), None,
         None, None, None,
         'full − griffin − merkle'),
    ]

    data_start = 3
    for i, (label, cc, pct, prove, verify, pb, notes) in enumerate(rows_data):
        r = data_start + i
        ws.cell(row=r, column=1, value=label)
        if cc is not None:
            write_val(ws, r, 2, round(cc) if cc else cc, '#,##0')
        if pct is not None:
            write_val(ws, r, 3, pct, '0.0%')
        if prove is not None:
            write_val(ws, r, 4, prove, '0.000')
        if verify is not None:
            write_val(ws, r, 5, verify, '0.000')
        if pb is not None:
            write_val(ws, r, 6, round(pb) if pb else pb, '#,##0')
        ws.cell(row=r, column=7, value=notes)

    style_rows(ws, data_start, data_start + len(rows_data) - 1, len(cols))
    for c in range(1, len(cols) + 1):
        ws.cell(row=data_start, column=c).font = BOLD_FONT

    # Per-size griffin rows
    size_row = data_start + len(rows_data) + 2
    ws.cell(row=size_row, column=1,
            value='Griffin-only: per hash input size'
            ).font = Font(name='Arial', bold=True, size=10, color='1F4E79')
    size_row += 1

    for hs_rec in suite_samples(samples, 'B5', 'griffin_only'):
        cfg = hs_rec.get('config', {})
        hs  = cfg.get('hash_input_size') or cfg.get('hash_inputs')
        if hs is None:
            continue
        cc = get_metrics(hs_rec, 'constraint_count') or 0
        ws.cell(row=size_row, column=1, value=f'  hash_input_size={hs}')
        write_val(ws, size_row, 2, cc, '#,##0')
        size_row += 1

    return ws

# ---------------------------------------------------------------------------
# Sheet 5: Backend Comparison (B4)
# ---------------------------------------------------------------------------

def build_sheet_backend(wb, samples):
    ws = wb.create_sheet()
    cols = ['Run#', 'Backend', 'constraint_count', 'prove_s', 'verify_s',
            'proof_bytes', 'proof_MiB', 'source_file']
    widths = [6, 12, 18, 12, 12, 14, 12, 30]
    setup_sheet(ws, 'Backend Comparison (B4)', cols, widths)

    all_recs = []
    for variant in ['aurora', 'fractal']:
        for s in suite_samples(samples, 'B4', variant):
            s['_backend'] = variant
            all_recs.append(s)

    if not all_recs:
        ws.cell(row=3, column=1, value='NO DATA').font = BOLD_FONT
        return ws

    all_recs.sort(key=lambda s: (s['_backend'], s.get('run', 0)))

    data_start = 3
    for i, s in enumerate(all_recs):
        r = data_start + i
        cc    = s.get('config', {}).get('constraint_count') or 0
        prove = (get_phases(s, 'prove_ms') or 0) / 1000
        vrfy  = (get_phases(s, 'verify_ms') or 0) / 1000
        pb    = get_metrics(s, 'proof_bytes') or 0

        write_val(ws, r, 1, s.get('run', i))
        write_val(ws, r, 2, s['_backend'])
        write_val(ws, r, 3, cc,    '#,##0')
        write_val(ws, r, 4, prove, '0.000')
        write_val(ws, r, 5, vrfy,  '0.000')
        write_val(ws, r, 6, pb,    '#,##0')
        write_val(ws, r, 7, pb / (1024**2) if pb else None, '0.00')
        write_val(ws, r, 8, Path(s.get('_source_file', '')).name)

    data_end = data_start + len(all_recs) - 1
    style_rows(ws, data_start, data_end, len(cols))

    # Per-backend summaries
    stat_row = data_end + 2
    for backend in ['aurora', 'fractal']:
        brecs = [s for s in all_recs if s['_backend'] == backend]
        if not brecs:
            continue
        prove_v  = [(get_phases(s, 'prove_ms') or 0) / 1000 for s in brecs]
        verify_v = [(get_phases(s, 'verify_ms') or 0) / 1000 for s in brecs]
        pb_v     = [get_metrics(s, 'proof_bytes') or 0 for s in brecs]
        cc_v     = [s.get('config', {}).get('constraint_count') or 0 for s in brecs]

        for label, fn in [
            ('MEAN',  statistics.mean),
            ('STDEV', lambda v: statistics.stdev(v) if len(v)>1 else 0),
        ]:
            style_summary(ws, stat_row, len(cols), f'{label} ({backend})')
            write_val(ws, stat_row, 3, fn(cc_v),     '#,##0')
            write_val(ws, stat_row, 4, fn(prove_v),  '0.000')
            write_val(ws, stat_row, 5, fn(verify_v), '0.000')
            write_val(ws, stat_row, 6, fn(pb_v),     '#,##0')
            write_val(ws, stat_row, 7, fn(pb_v) / (1024**2), '0.00')
            stat_row += 1

    # Delta row
    aurora_recs  = [s for s in all_recs if s['_backend'] == 'aurora']
    fractal_recs = [s for s in all_recs if s['_backend'] == 'fractal']
    if aurora_recs and fractal_recs:
        a_prove  = statistics.mean([(get_phases(s,'prove_ms') or 0)/1000 for s in aurora_recs])
        f_prove  = statistics.mean([(get_phases(s,'prove_ms') or 0)/1000 for s in fractal_recs])
        a_verify = statistics.mean([(get_phases(s,'verify_ms') or 0)/1000 for s in aurora_recs])
        f_verify = statistics.mean([(get_phases(s,'verify_ms') or 0)/1000 for s in fractal_recs])
        a_pb = statistics.mean([get_metrics(s,'proof_bytes') or 0 for s in aurora_recs])
        f_pb = statistics.mean([get_metrics(s,'proof_bytes') or 0 for s in fractal_recs])

        style_summary(ws, stat_row, len(cols), 'Δ% Fractal vs Aurora')
        write_val(ws, stat_row, 4, (f_prove - a_prove) / a_prove if a_prove else None,  '0.0%')
        write_val(ws, stat_row, 5, (f_verify - a_verify) / a_verify if a_verify else None, '0.0%')
        write_val(ws, stat_row, 6, (f_pb - a_pb) / a_pb if a_pb else None, '0.0%')

    return ws

# ---------------------------------------------------------------------------
# Sheet 6: D1 Update Churn (static / qualitative)
# ---------------------------------------------------------------------------

def build_sheet_churn(wb):
    ws = wb.create_sheet()
    cols = ['Update Event', 'Aurora: Artifacts',
            'Aurora: Effort', 'Aurora: BW (bytes)',
            'zkVM: Artifacts', 'zkVM: Effort', 'zkVM: BW (bytes)',
            'BW Ratio (A/Z)']
    widths = [22, 30, 16, 18, 30, 16, 18, 16]
    setup_sheet(ws, 'D1 Update Churn', cols, widths)

    churn_rows = [
        ('Policy update',
         'R1CS + proving key + verification key', 'Hours–days', 33_554_432,
         'Guest binary (recompile)', 'Minutes', 32),
        ('Algorithm rotation',
         'All circuit artifacts + keys', 'Weeks–months', 33_554_432,
         'Guest binary + host logic', 'Hours–days', 32),
        ('Schema extension',
         'R1CS + keys + Noir source', 'Days–weeks', 33_554_432,
         'Guest binary', 'Hours', 32),
        ('Parameter refresh',
         'Setup params + keys', 'Hours', 33_554_432,
         'Config update', 'Minutes', 32),
    ]

    for i, row_data in enumerate(churn_rows):
        r = 3 + i
        event, a_art, a_eff, a_bw, z_art, z_eff, z_bw = row_data
        ws.cell(row=r, column=1, value=event)
        ws.cell(row=r, column=2, value=a_art)
        ws.cell(row=r, column=3, value=a_eff)
        write_val(ws, r, 4, a_bw, '#,##0')
        ws.cell(row=r, column=5, value=z_art)
        ws.cell(row=r, column=6, value=z_eff)
        write_val(ws, r, 7, z_bw, '#,##0')
        write_val(ws, r, 8, a_bw / z_bw if z_bw else None, '#,##0×')

    style_rows(ws, 3, 2 + len(churn_rows), len(cols))
    return ws

# ---------------------------------------------------------------------------
# Sheet 7: Cross-Backend Summary
# ---------------------------------------------------------------------------

def build_sheet_summary(wb, samples):
    ws = wb.create_sheet()
    cols = ['Metric', 'Aurora (M5 Pro)', 'RISC Zero (dev mode)', 'Notes']
    widths = [32, 22, 22, 50]
    setup_sheet(ws, 'Cross-Backend Summary', cols, widths)

    # Pull Aurora PP2 numbers
    pp2 = suite_samples(samples, 'B7', 'pp2_aurora')
    prove_m, prove_s  = mean_std([(get_phases(s,'prove_ms') or 0)/1000 for s in pp2]) if pp2 else (None,None)
    verify_m, verify_s= mean_std([(get_phases(s,'verify_ms') or 0)/1000 for s in pp2]) if pp2 else (None,None)
    pb_m, _           = mean_std([get_metrics(s,'proof_bytes') or 0 for s in pp2]) if pp2 else (None,None)
    cc_pp2            = pp2[0].get('metrics',{}).get('constraint_count') if pp2 else None

    rows_data = [
        ('Constraint count / Trace cycles',
         f'{cc_pp2:,}' if cc_pp2 else '—',
         '~3.8 B cycles (k=1, s=1, m=16)',
         'R1CS constraints vs RISC-V execution cycles'),
        ('Prove time (s)',
         f'{prove_m:.1f} ± {prove_s:.1f}' if prove_m else '—',
         '~22 s (dev mode, cycle count proxy)',
         'Wall-clock proving time; Aurora on M5 Pro'),
        ('Verify time (s)',
         f'{verify_m:.1f} ± {verify_s:.1f}' if verify_m else '—',
         '< 0.01 s (dev mode)',
         'Aurora verifier slow due to IOP model; zkVM verifier is O(1)'),
        ('Proof size',
         f'{pb_m/1024**2:.1f} MiB IOP transcript' if pb_m else '—',
         '~215 KiB (STARK proof)',
         'Aurora = full IOP transcript; zkVM = compact STARK'),
        ('Verification key size',
         '~32 MB (circuit-specific)', '32 B (ImageID)',
         'Aurora VK = full R1CS encoding'),
        ('Policy update churn',
         'Hours–days (R1CS recompile)', 'Minutes (binary recompile)',
         'D1 dimension'),
        ('Field alignment (Griffin p)',
         '1× (native Fp²)',
         '4× (BabyBear 32-bit limbs)',
         'p = 2^127−1 on BabyBear p = 2^31−2^27+1'),
        ('Precompile paradox',
         'Griffin: 304 constraints/round',
         'SHA-256: 1 cycle (precompile)',
         'Optimal hash is opposite for each backend'),
        ('Agility tax',
         'High (circuit tied to hash)',
         'Low (hash is a library call)',
         'Algorithm rotation cost'),
    ]

    for i, (metric, aurora_val, zkvm_val, notes) in enumerate(rows_data):
        r = 3 + i
        ws.cell(row=r, column=1, value=metric)
        ws.cell(row=r, column=2, value=aurora_val)
        ws.cell(row=r, column=3, value=zkvm_val)
        ws.cell(row=r, column=4, value=notes)

    style_rows(ws, 3, 2 + len(rows_data), len(cols))
    return ws

# ---------------------------------------------------------------------------
# Sheet 8: Per-run raw data (audit trail)
# ---------------------------------------------------------------------------

def build_sheet_raw(wb, samples):
    ws = wb.create_sheet()
    cols = ['suite', 'variant', 'config_key', 'run', 'is_warmup',
            'wall_ms', 'prove_ms', 'verify_ms',
            'constraint_count', 'proof_bytes', 'source_file']
    widths = [8, 20, 28, 6, 10, 14, 14, 14, 16, 14, 30]
    setup_sheet(ws, 'Raw Data', cols, widths)

    all_s = sorted(samples, key=lambda s: (
        s.get('suite', ''), s.get('variant', ''), s.get('config_key', ''), s.get('run', 0)
    ))

    for i, s in enumerate(all_s):
        r = 3 + i
        write_val(ws, r, 1,  s.get('suite', ''))
        write_val(ws, r, 2,  s.get('variant', ''))
        write_val(ws, r, 3,  s.get('config_key', ''))
        write_val(ws, r, 4,  s.get('run', ''))
        write_val(ws, r, 5,  str(s.get('is_warmup', '')))
        write_val(ws, r, 6,  s.get('wall_ms'),  '0.0')
        write_val(ws, r, 7,  get_phases(s, 'prove_ms'),  '0.0')
        write_val(ws, r, 8,  get_phases(s, 'verify_ms'), '0.0')
        write_val(ws, r, 9,  get_metrics(s, 'constraint_count'), '#,##0')
        write_val(ws, r, 10, get_metrics(s, 'proof_bytes'), '#,##0')
        write_val(ws, r, 11, Path(s.get('_source_file', '')).name)

    style_rows(ws, 3, 2 + len(all_s), len(cols))
    return ws

# ---------------------------------------------------------------------------
# Choose canonical input files
# ---------------------------------------------------------------------------

def choose_jsonl_files(cli_files=None):
    """
    Return ordered list of JSONL files to load.
    Priority: CLI --file args > most-recent bench_TIMESTAMP.jsonl files.
    The canonical M2 files are always included as fallback.
    """
    if cli_files:
        return [Path(f) for f in cli_files if Path(f).exists()]

    all_bench = sorted(RESULTS_DIR.glob('bench_*.jsonl'))
    if not all_bench:
        print("WARNING: no bench_*.jsonl files found in results/", file=sys.stderr)
        return []

    # Use all files — caller can filter by suite/variant using source_file metadata
    return all_bench

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Build Excel workbook from JSONL bench results')
    parser.add_argument('--output', default=str(RESULTS_DIR / 'thesis_benchmark_analysis.xlsx'))
    parser.add_argument('--file', nargs='+', help='Specific JSONL files to load (default: all bench_*.jsonl)')
    args = parser.parse_args()

    if not HAS_OPENPYXL:
        sys.exit(1)

    jsonl_files = choose_jsonl_files(args.file)
    if not jsonl_files:
        print("No input files found.", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {len(jsonl_files)} JSONL file(s)...")
    samples, summaries, headers = collect_records(jsonl_files)
    print(f"  {len(samples)} sample records, {len(summaries)} summary records")

    # Print suite/variant breakdown
    breakdown = defaultdict(set)
    for s in samples:
        breakdown[s.get('suite', '?')].add(s.get('variant', '?'))
    for suite, variants in sorted(breakdown.items()):
        n = len([s for s in samples if s.get('suite') == suite])
        print(f"  {suite}: {sorted(variants)}  ({n} records)")

    # Filter to canonical (real libiop) records
    canon = canonical(samples)
    toy_filtered = len(samples) - len(canon)
    if toy_filtered:
        print(f"  Filtered out {toy_filtered} toy-aurora / non-canonical records")

    wb = Workbook()

    # Choose representative header (M5 Pro if available, else first)
    m5_headers = [h for h in headers if 'M5' in h.get('cpu', '') or '1776' in h.get('_source_file', '')]
    best_header = m5_headers or headers

    build_sheet_pp2(wb, canon, best_header)
    build_sheet_pp3(wb, canon)
    build_sheet_scaling(wb, canon, summaries)
    build_sheet_griffin(wb, canon)
    build_sheet_backend(wb, canon)
    build_sheet_churn(wb)
    build_sheet_summary(wb, canon)
    build_sheet_raw(wb, canon)

    output = Path(args.output)
    wb.save(str(output))
    print(f"\nSaved → {output}")
    print("Sheets: " + ", ".join(ws.title for ws in wb.worksheets))


if __name__ == '__main__':
    main()
